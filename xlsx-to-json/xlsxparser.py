from openpyxl import load_workbook
import json
from tkinter import filedialog
import os
from re import match

print('Welcome to xlsx-to-json 1.0!')
print('----------------------------')
print('This small script will read an excel file and convert it into and json file containing a list of objects, which all represent one line of you excel file, from the chosen columns.')

while True:
    while True:
        msg = input('Open File (Y/N)>')
        if msg.lower() == 'y':
            f = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')], initialdir=os.getcwd())
            if not f or not os.path.exists(f) or not os.path.getsize(f):
                continue
            else: break
        elif msg.lower() == 'n': quit()
    try: 
        wb = load_workbook(filename=f)
    except:
        print('Invalid File')
        continue
    print('Avaliable Sheets:')
    print(wb.sheetnames)
    while True:
        inp = input('Which Sheet do you want to read from? (type "exit" to quit)>')
        if inp in wb.sheetnames: break
        elif inp.lower() == 'exit': quit()
    ws = wb[inp]
    while True:
        columns = input('Enter column names seperated by | (capitalized). Example: A|C|D|BD>')
        mat1 = match('^[ABCDEFGHIJKLMNOPQRSTUVWXYZ|]*$', columns)
        if mat1 is None:
            print('Wrong column names')
            continue
        elif columns.lower() == 'exit': quit()
        else: 
            ids = columns.split('|')
        rng = int(input('Enter number of rows you want to parse. (1000 max, aborts on empty line)>'))
        if rng > 1000 or rng < 1:
            print('Invalid row number')
            continue
        attributes = input('Enter attribute names for the json object seperated by |. (You need as many attributes as column names)>')
        attrs = attributes.split('|')
        if not len(attrs) == len(ids):
            print('Column and attribute number not matching!')
            continue
        elif columns.lower() == 'exit': quit()
        else: break
    print('Processing...')
    data = {'content':[]}
    cells = dict()
    for i in range(1, rng+1):
        for a,c in enumerate(ids):
            cell = ws[c+str(i)].value
            if not cell == '' and not cell == None:
                cells[attrs[a]] = cell
            else:
                cells[attrs[a]] = ''
        if not any(item for item in cells.values()):
            break
        data['content'].append(cells)
        cells = dict()
    while True:
        print('Select save folder and filename:')
        f2n = filedialog.asksaveasfilename(filetypes=[('Json File', '*.json')], initialfile='output', initialdir=os.getcwd())
        if not f2n:
            continue
        else: break
    try: 
        if not f2n[-5:] == '.json': f2n = f2n + '.json'
        with open(f2n, 'w') as f2:
            json.dump(data, f2)
    except Exception as e:
        print('ERROR!')
        print(e)
        msg3 = input('Do you want to retry? (Y/N)>')
        if msg3.lower() == 'y':
            continue
        else: quit()
    print('Conversion successful!')
    quit()